
print(40 * "-")
print("Downloading packages ...")


# Standard library
import io
import itertools
import os
import random
import sys
import time
import warnings
from abc import ABC, abstractmethod

start_package = time.time()

# External libraries
import argparse
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

from catboost import CatBoostRegressor
from group_lasso import GroupLasso
from more_itertools import flatten
from scipy.linalg import svd
from scipy.stats import norm, kurtosis, skew
from sklearn.cross_decomposition import PLSRegression
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import LinearRegression, HuberRegressor, ElasticNet
from sklearn.model_selection import ParameterGrid, ParameterSampler
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import KBinsDiscretizer, MinMaxScaler
from skopt import Optimizer
from skopt.callbacks import DeltaYStopper
from skopt.space import Real, Integer, Categorical
from skopt.utils import create_result
from tensorflow import add, reduce_sum, divide, subtract, square
from tensorflow.keras.callbacks import EarlyStopping
from tensorflow.keras.layers import BatchNormalization, Dense, Dropout, Input
from tensorflow.keras.models import Sequential
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.regularizers import l1
from tensorflow.random import set_seed
from xgboost import XGBRFRegressor, XGBRegressor


end_package = time.time()

print("Packages downloaded succesfully !!!")
print("It took {:.2f} seconds".format(end_package - start_package))
print(40 * "-", "\n", "\n")


%matplotlib
warnings.filterwarnings('ignore')




##########################
# COMMAND LINE INTERACTION
##########################

parser = argparse.ArgumentParser(
    description="Run specified models and store the results in an excel file called results_model.xlsx",
    formatter_class=argparse.ArgumentDefaultsHelpFormatter
)

models_description = {
    "ols": "Ordinary Least Squares",
    "ols3": "Ordinary Least Squares with 3 factors",
    "pls": "Partial Least Squares Regression",
    "pcr": "Principal Component Regression",
    "enet": "Elastic Net Regression with L1 penalization",
    "glm": "Generalized Linear Model with Group Lasso regularization",
    "rf": "Random Forest",
    "gb": "Gradient Boosting",
    "nn1": "Neural Network with 1 hidden layer",
    "nn2": "Neural Network with 2 hidden layers",
    "nn3": "Neural Network with 3 hidden layers",
    "nn4": "Neural Network with 4 hidden layers",
    "nn5": "Neural Network with 5 hidden layers"
}

# Add all the arguments for the command line 
for name, description in models_description.items():
    parser.add_argument(f"--run_{name}", action="store_true", help=f"Specify whether to run the {description} model")







##################
# DATE IMPORTATION
################## 

print(40 * "-")
print("Importing data ...")
start_import = time.time()

X_macro = pd.read_csv('Macroeconomic_series.txt', delimiter='\t')

# Initialize empty dataframes for y and X_ind
y = pd.DataFrame()
X_ind = pd.DataFrame()

# List all files in the "Stocks Prices - Returns" and "Stocks Features" directories
returns_dir = 'Stocks Prices - Returns'
features_dir = 'Stocks Features'
returns_files = os.listdir(returns_dir)
features_files = os.listdir(features_dir)

# Loop through the files and import them
for return_file, feature_file in zip(returns_files, features_files):
    try:
        y_int = pd.read_csv(os.path.join(returns_dir, return_file), delimiter='\t')
        X_ind_int = pd.read_csv(os.path.join(features_dir, feature_file), delimiter='\t')
        y = pd.concat([y, y_int], axis=0)
        X_ind = pd.concat([X_ind, X_ind_int], axis=0)
    except:
        continue    
    
end_import = time.time()
print( "Data imported succesfully !!!")
print("It took {:.2f} seconds".format(end_import - start_import))
print(40 * "-", "\n", "\n")


    
##################
# DATE EXPLORATION
##################   


# Individual characteristics X_ind 
print(X_ind.head())
print(X_ind.tail())
print(X_ind.shape)
print(X_ind.info())


# Macroeconomic variables X_macr
print(X_macro.head())
print(X_macro.tail())
print(X_macro.shape)
print(X_macro.info())


# Dependent variable y
print(y.head())
print(y.tail())
print(y.shape)
print(y.info())




#####################
# DATE PRE-PROCESSING
#####################

print(40 * "-")
print("Pre-processing the data ...")
start_prep = time.time()


# The dependent variable y:    
y = y[["PERMNO", "date", "RET"]]
y["date"] = y["date"].astype(str)
y['date'] = y['date'].apply(lambda x: x[:-2])
y["date"] = pd.to_datetime(y["date"], format = "%Y%m") + pd.offsets.MonthEnd(1)
y.rename(columns = {"PERMNO": "permno"}, inplace = True)
y = y.set_index(["date", "permno"])
y.sort_index(level = 0, ascending = True, inplace = True)
y['RET'] = pd.to_numeric(y['RET'], errors='coerce')

#print(y.isna().sum())



# The individual stock's characteristics  X_ind:
X_ind.rename(columns={'DATE': 'date'}, inplace = True)    
X_ind["date"] = X_ind["date"].astype(str)
X_ind['date'] = X_ind['date'].apply(lambda x: x[:-2])
X_ind["date"] = pd.to_datetime(X_ind["date"], format = "%Y%m") + pd.offsets.MonthEnd(1)
X_ind = X_ind.set_index(["date", "permno"])
X_ind.sort_index(level = 0, ascending = True, inplace = True)
X_ind = X_ind.apply(pd.to_numeric, errors = "coerce")

df = y.join(X_ind, how = "inner")
df.dropna(how='all', inplace = True)



# The macroeconimic variables X_macr:
X_macro = X_macro.rename(columns={'yyyymm': 'date'})
X_macro["date"] = X_macro["date"].astype(str)
X_macro["date"] = pd.to_datetime(X_macro["date"], format = "%Y%m") + pd.offsets.MonthEnd(1)
X_macro.set_index("date", inplace = True)
X_macro.sort_index(level = 0, ascending = True, inplace = True)


# Create the excess return and merge it with X_ind
rf = X_macro["Rfree"]
df_merged = pd.merge(df, rf, on = "date")
df_merged.index = df.index
df_merged["Excess Return"] = df_merged["RET"] - df_merged["Rfree"]
X_macro.drop("Rfree", axis = 1, inplace = True)
df_merged.drop(["RET", "Rfree"], axis = 1, inplace=  True)
merged_df = pd.merge(df_merged.reset_index(level = ["permno"]), X_macro, on='date')

df = merged_df.set_index([merged_df.index, 'permno'])





# Impute categorical variables:
sic2 = df[["sic2"]]
df.drop("sic2", axis = 1, inplace = True)

def impute_mode(x):
    mode_series = x.mode()
    if mode_series.empty:
        return 'Unknown'
    else:
        return mode_series.iloc[0]


sic2_imputed = sic2.groupby(level = 1).apply(lambda x: x.fillna(impute_mode(x)))
sic2_imputed = sic2_imputed.droplevel(level = 0)
sic2_imputed.sort_index(level = "date", ascending = True, inplace = True)
sic2_imputed = sic2_imputed.astype("category")



# Impute numerical variables:
# dropping all rows of NA values from the return variable
na_rows = df[df["Excess Return"].isna()].index
df.drop(na_rows, axis = 0, inplace = True)
sic2_imputed.drop(na_rows, axis = 0, inplace = True)
df_imputed = df.groupby(level='date').apply(lambda x: x.fillna(x.median())).droplevel(0)
df_imputed.fillna(0, inplace = True)




# Lag the regressors according to their lagging periods 
table = pd.read_excel('lagg_names.xlsx')
df_imputed.rename(columns = {"ep_x": "ep"}, inplace = True)
names_ind = df_imputed.loc[:, : "zerotrade"].columns.tolist()
names_macro = df_imputed.loc[:, "b/m": ].columns.tolist()

annual_table = table[table['Frequency'] == 'Annual']
monthly_table = table[table['Frequency'] == 'Monthly']
quarterly_table = table[table['Frequency'] == 'Quarterly']

annual_acronyms = annual_table['Acronym'].tolist()
monthly_acronyms = monthly_table['Acronym'].tolist()
quarterly_acronyms = quarterly_table['Acronym'].tolist()


df_imputed[annual_acronyms] = df_imputed[annual_acronyms].groupby(level=1).transform(lambda x: x.shift(6))
df_imputed[monthly_acronyms + names_macro] = df_imputed[monthly_acronyms + names_macro].groupby(level=1).transform(lambda x: x.shift(1))
df_imputed[quarterly_acronyms] = df_imputed[quarterly_acronyms].groupby(level=1).transform(lambda x: x.shift(4))

# Filter NA values in both df_imputed and sic2_imputed
na_rows = df_imputed[df_imputed.isna().any(axis = 1)].index
df_imputed.drop(na_rows, axis = 0, inplace = True)
sic2_imputed.drop(na_rows, axis = 0, inplace = True)




# Cross-sectional ranking of individual characteristics and normalization betwen (-1,1)
ranked_charac = df_imputed.loc[:, :"zerotrade"].groupby(level=0).apply(lambda x: x.rank()).droplevel(0)
scaler = MinMaxScaler((-1, 1))
normalized_charac = ranked_charac.groupby(level=0).apply(lambda x: pd.DataFrame(scaler.fit_transform(x), index=x.index, columns=x.columns)).droplevel(0)
ranked_df = pd.concat([normalized_charac, df_imputed.loc[:, "Excess Return":]], axis = 1)




# Create interaction effects between X_ind and X_macro
ranked_df.rename(columns = {"ep_y": "ep_macro"}, inplace = True)
names_macro = ranked_df.loc[:, "b/m":].columns.to_list()
names_ind = ranked_df.loc[:, :"zerotrade"].columns.to_list()
combinations = [' * '.join(map(str, combo)) for combo in itertools.product(names_ind, names_macro)]

df_indiv = ranked_df.loc[:, :"zerotrade"]
df_new = df_indiv

for name in names_macro:
      
    inter = df_indiv.multiply(ranked_df[name], axis = 0)
    df_new = pd.concat([df_new, inter], axis = 1)

names = names_ind + combinations
df_new.columns = names




# One hot-encoding the SIC code
dummies = pd.get_dummies(sic2_imputed, prefix = "sic2", drop_first = True, dtype = int)

def clean_column_name(col_name):
    return col_name.replace(".0", "")

dummies.columns = [clean_column_name(col_name) for col_name in dummies.columns]

df_new = pd.concat([ranked_df["Excess Return"], df_new, dummies], axis = 1)

# Delete unnecessary data frames
del(df, df_imputed, df_indiv, df_merged, dummies, inter, merged_df, na_rows)
del(normalized_charac, ranked_charac, ranked_df, rf, sic2, sic2_imputed, X_ind)
del(X_ind_int, X_macro, y, y_int)


end_prep = time.time()
print("Data pre-processed succesfully !!!")
print("It took {:.2f} seconds".format(end_prep - start_prep))
print(40 * "-", "\n", "\n")






####################
# MODEL CONSTRUCTION
####################


class Model(ABC):
    """
    An abstract base class for machine learning models in the context of replicating the models of the paper "Empirical Asset Pricing via Machine Learning" by Gu et al. (2020).
    This class serves as a template for creating custom machine learning models as described in the paper.
    It includes methods for setting and getting parameters, as well as abstract methods for fitting, predicting, and scoring that must be implemented in subclasses.
    
    Attributes:
        hyperparams (dict): A dictionary of model hyperparameters.
        
    Methods
    -------
    r2_oos(actual, pred)
        Static method for computing the Pooled out-of-sample R-Squared by assuming an expected excess reurn of 0
    set_params(params)
        Sets the model parameters.
    get_params()
        Returns the model parameters.
    fit(X_train, y_train)
        Abstract method for fitting the model on the given training data.
    predict(X_test, y_test)
        Abstract method for predicting the output for the given input data.
    score(X_test, y_test)
        Abstract method for computing the performance score on the test data.
    """
    
    
    def __init__(self):
        self.hyperparams = {}
        
    @staticmethod    
    def r2_oos(actual, pred):
        actual, pred = np.array(actual).flatten(), np.array(pred).flatten()
        pred = np.where(pred < 0, 0, pred)
        return 1 - (np.dot((actual - pred), (actual - pred))) / (np.dot(actual, actual))
    
    @staticmethod
    def loss_mse(actual, pred):
        actual, pred = np.array(actual).flatten(), np.array(pred).flatten()
        resid = actual - pred
        return np.sum(resid**2) / len(actual)
    
    @staticmethod
    def loss_huber(actual, pred, epsilon):
        actual, pred = np.array(actual).flatten(), np.array(pred).flatten()
        abs_resid = np.abs(actual - pred)
        huber_loss = np.where(abs_resid <= epsilon, abs_resid**2, 2 * epsilon * abs_resid - epsilon**2)
        return np.mean(huber_loss)
    
    @staticmethod
    def grad_mse(X, y, theta):
        X = np.array(X)
        N = len(y)
        y = np.array(y).reshape((N,1))
        theta = np.array(theta).reshape((X.shape[1],1))
        return (X.T @ (y - X @ theta))/N
    
    @staticmethod
    def grad_huber(X, y, theta, epsilon):
        K = X.shape[1]
        X = np.array(X)
        N = len(y)
        y = np.array(y).reshape((N,1))
        theta = np.array(theta).reshape((K,1))
        residual = y - X @ theta
        grad = np.zeros((K,1))
        for i, r in enumerate(residual):
            if np.abs(r) <= epsilon:
                grad += X[i].reshape(-1, 1) * r
            elif r > epsilon:
                grad += 2 * epsilon * X[i].reshape(-1, 1)
            else:
                grad -= 2 * epsilon * X[i].reshape(-1, 1)
        return grad / N

    def set_params(self, params):
        self.params = params
    
    def get_params(self):        
        return self.params
    
    @abstractmethod
    def fit(self, X_train, y_train):   
        """
        Fit a model to the training data.

        Parameters
        ----------
        X_train : array-like of shape (n_samples, n_features)
            The training input samples.

        y_train : array-like of shape (n_samples,)
            The target values.

        Returns
        -------
        self : object
            Returns the instance itself.
        """

        pass
    
    @abstractmethod
    def predict(self, X_test, y_test):    
        """
        Predict the target values for the test data.

        Parameters
        ----------
        X_test : array-like of shape (n_samples, n_features)
            The input samples.

        Returns
        -------
        y_pred : array-like of shape (n_samples,)
            The predicted target values for the input samples.
        """

        pass
    
    @abstractmethod
    def score(self, X_test, y_test):
        """
        Evaluate the performance of the model on the test data.

        Parameters
        ----------
        X_test : array-like of shape (n_samples, n_features)
            The input samples.

        y_test : array-like of shape (n_samples,)
            The target values.

        Returns
        -------
        float
            The out-of-sample R2of the model on the test data.
        """

        pass
    


class OrdinaryLeastSquares(Model):
    """
    Ordinary Least Squares (OLS) linear regression model, with options for mean squared error (MSE) or Huber loss functions.

    Parameters
    ----------
    cost_function : str, optional (default="huber")
        The loss function to be used for the OLS model, either "mse" or "huber".

    huber_quantile : float, optional (default=99.9)
        The quantile to be used for calculating epsilon in the Huber loss function.

    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.

    set_params(**params)
        Set the parameters of the model.

    compute_epsilon(X_train, y_train, huber_quantile)
        Compute the epsilon value for the Huber loss function.

    build_model()
        Build the sklearn model for the OLS model with the specified cost function.

    fit(X_train, y_train)
        Fit the model to the training data.

    predict(X_test)
        Predict the target values for the test data.

    score(X_test, y_test)
        Evaluate the model's performance on the test data.
    """

    
    def __init__(self, cost_function="huber", huber_quantile=99.9):
        self.cost_function = cost_function
        self.huber_quantile = None
        
        if self.cost_function not in ["mse", "huber"]:
            raise ValueError("Invalid value for parameter 'cost_function', specify either 'mse' or 'huber'")
            
        if self.cost_function == "huber":
            self.huber_quantile = huber_quantile
            
        self._build_model()
            
    def get_params(self, deep=True):
        mse_params = {
            "cost_function": self.cost_function,
        }
        
        if self.cost_function == "mse":
            return mse_params
        else:
            huber_params = {"huber_quantile": self.huber_quantile}
            return {**mse_params, **huber_params}

    def set_params(self, **params):
        rebuild_model = False
        for param, value in params.items():
            
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{param}' for estimator {self.__class__.__name__}") 
        if rebuild_model:
            self._build_model()
    
    def __compute_epsilon(self, X_train, y_train, huber_quantile):
        """
        Private method that computes the value of the epsilon parameter for the Huber loss function.
        
        It does it by an ols model and set the epsilon value as the specified qauntile of the model's residuals

        Parameters
        ----------
        X_train : array-like of shape (n_samples, n_features)
            The training input samples.

        y_train : array-like of shape (n_samples,)
            The target values.

        huber_quantile : float
            The quantile used to compute the epsilon parameter.

        Returns
        -------
        float
            The computed value of the epsilon parameter.
        """

        lr = LinearRegression()
        lr.fit(X_train, y_train)
        y_pred = lr.predict(X_train)
        abs_resid = np.abs(y_train - y_pred)
        epsilon = max(np.percentile(abs_resid, self.huber_quantile), 1)
        return epsilon

    def _build_model(self):
        """
        Protected method that creates a protected model attribute based on the selected cost function.
        """

        if self.cost_function == "huber":
            self.model = HuberRegressor()
        else:
            self.model = LinearRegression()
            
    def fit(self, X_train, y_train):
        y_train = y_train.ravel()
        
        if self.cost_function == "huber":
            epsilon = self.__compute_epsilon(X_train, y_train, self.huber_quantile)
            self.model.set_params(epsilon = epsilon)
            self.model.fit(X_train, y_train)
        else:
            self.model.fit(X_train, y_train)
            
        return self
               
    def predict(self, X_test):
        return self.model.predict(X_test).flatten()

    def score(self, X_test, y_test):
        y_pred = self.predict(X_test)
        r2_oos = self.r2_oos(y_test, y_pred)
        return r2_oos

    


class PrincipalComponentRegression(OrdinaryLeastSquares):
    """
    Principal Component Regression (PCR) model that uses ordinary least squares regression on the transformed feature matrix.

    Parameters
    ----------
    n_components : int, float, str or None, optional (default=None)
        The number of principal components to retain. If None, all components are kept.

    args : list
        Optional positional arguments to pass to the superclass constructor.

    kwargs : dict
        Optional keyword arguments to pass to the superclass constructor.


    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.

    set_params(**params)
        Set the parameters of the model.

    build_model()
        Build the PCR model by chaining the PCA transformer and the linear regression model.

    fit(X_train, y_train)
        Fit the model to the training data.

    predict(X_test)
        Predict the target values for the test data.

    score(X_test, y_test)
        Compute the out-of-sample R-squared score for the test data.
    """
    
    def __init__(self,
                 n_components=None,
                 *args,
                 **kwargs
                 ):
        self.n_components = n_components
        super().__init__(*args, **kwargs)
        self.cost_function = "mse"
        self._build_model()
    
    def get_params(self, deep=True):
        parent_params = super().get_params(deep)
        child_params = {"n_components": self.n_components}
        params = {**parent_params, **child_params}
        return params

    def set_params(self, **params):         
        rebuild_model = False
        for param, value in params.items():
            
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{param}' for estimator {self.__class__.__name__}") 
                
        if rebuild_model:
            self._build_model()
        
    def build_model(self):
        """
        Build the PCR model as a Pipeline object by chaining the PCA transformer and the linear regression model.
        """

        self.pca = PCA(n_components=self.n_components)
        super()._build_model()
        self.lr = self.model
        self.model = Pipeline([("pca", self.pca), ("lr", self.lr)])
        
    def fit(self, X_train, y_train):
        self.model.fit(X_train, y_train)
        return self
            
    def predict(self, X_test):
        return self.model.predict(X_test).flatten()
        
    def score(self, X_test, y_test):        
        y_pred = self.predict(X_test)
        r2_oos = self.r2_oos(y_test, y_pred)
        return r2_oos
            
        


class PartialLeastSquaresRegression(OrdinaryLeastSquares):
    """
    Partial Least Squares (PLS) regression model, with MSE as the cost function.

    Parameters
    ----------
    n_components : int or None, optional (default=None)
        The number of PLS components to use in the model. If None, the optimal number of components will be computed.

    cost_function : str, optional (default="mse")
        The cost function to be used for the PLS model.

    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.

    set_params(**params)
        Set the parameters of the model.

    build_model()
        Build the PLS regression model based on the selected number of components.

    fit(X_train, y_train)
        Fit the model to the training data.

    predict(X_test)
        Predict the target values for the test data.

    score(X_test, y_test)
        Evaluate the model's performance on the test data.
    """
    
    def __init__(self,
                 n_components=None,
                 *args,
                 **kwargs
                 ):
        self.n_components = n_components
        super().__init__(*args, **kwargs)
        self.build_model()
        self.cost_function = "mse"
    
    def get_params(self, deep=True):
        parent_params = super().get_params(deep)
        child_params = {"n_components": self.n_components}
        params = {**parent_params, **child_params}
        return params

    def set_params(self, **params):        
        rebuild_model = False
        for param, value in params.items():
            
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{param}' for estimator {self.__class__.__name__}") 
                
        if rebuild_model:
            self.build_model()
            
    def build_model(self):
        if self.n_components is None:
            self.model = PLSRegression()
        else:
            self.model = PLSRegression(n_components=self.n_components)
        
    def fit(self, X_train, y_train):
        self.model.fit(X_train, y_train)
        return self
            
    def predict(self, X_test):
        return self.model.predict(X_test).flatten()

    def score(self, X_test, y_test):
        y_pred = self.predict(X_test)
        r2_oos = self.r2_oos(y_test, y_pred)
        return r2_oos




class ElasticNetRegression(OrdinaryLeastSquares):
    """
    Elastic Net regression model that incorporate built-in and customized fitting process.
    
    This class accomodates for both mse and huber attribute by providing the appropriate fitting method.
    
    In the case where cost_function = 'huber' the approximal gradient algorithm is performed
    

    Parameters
    ----------
    lmd : float, optional (default=0.001)
        The L1 regularization parameter, which controls the sparsity of the model.

    rho : float, optional (default=0.5)
        The mixing parameter between L1 and L2 regularization. A value of 1 corresponds to Lasso regression,
        and a value of 0 corresponds to Ridge regression.

    max_iter : int, optional (default=1000)
        The maximum number of iterations for the coordinate descent algorithm.

    tol : float, optional (default=1e-4)
        The tolerance for convergence of the coordinate descent algorithm.

    method : str, optional (default="sklearn")
        The method for solving the Elastic Net problem. "sklearn" uses the scikit-learn implementation,
        while "coordinate_descent" uses a custom implementation based on coordinate descent.

    random_state : int, optional (default=None)
        The random seed for initializing the parameter estimates.

    verbose : bool, optional (default=False)
        Whether to print information about the fitting process.

    sleep : float, optional (default=None)
        The number of seconds to sleep after printing each iteration.

    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.

    set_params(**params)
        Set the parameters of the model.

    soft_threshold(x, mu)
        Apply the soft threshold function to a vector x with parameter mu.

    proximal(theta, lmd, rho, gamma)
        Apply the proximal operator to a vector theta with parameters lmd, rho, and gamma.

    __compute_L(X_train)
        Private method that computes the Lipschitz constant for the coordinate descent algorithm.

    build_model()
        Build the Elastic Net model.

    fit(X_train, y_train)
        Fit the model to the training data.

    predict(X)
        Predict the target values for the test data.

    score(X, y)
        Evaluate the model's performance on the test data.
    """

    
    def __init__(
        self,
        lmd=0.001,
        rho=0.5,
        max_iter=1000,
        tol=1e-4,
        method="sklearn",
        random_state=None,
        verbose=False,
        sleep=None,
        *args,
        **kwargs
        ):
        self.lmd = lmd
        self.rho = rho
        self.max_iter = max_iter
        self.tol = tol
        self.random_state = random_state
        self.verbose = verbose
        self.sleep = sleep
        self.method = method
        
        super().__init__(*args, **kwargs)
    
        if self.method == "sklearn" and self.cost_function == "huber":
            raise ValueError("Sklearn incompatible with huber cost function, please choose 'mse'")

        if method == "sklearn":
            self.build_model()

    def soft_threshold(self, x, mu):
        x = np.where(np.abs(x) <= mu, 0, x)
        x = np.where((np.abs(x) > mu) & (x > 0), x - mu, x)
        x = np.where((np.abs(x) > mu) & (x < 0), x + mu, x)
        return x
    
    def proximal(self, theta, lmd, rho, gamma):
        return (1 / (1 + lmd * gamma * rho)) * self.soft_threshold(theta, (1 - rho) * gamma * lmd)

    def get_params(self, deep = True):
        parent_params = super().get_params(deep)
        child_params = {
            "lmd" : self.lmd,
            "rho": self.rho,
            "max_iter": self.max_iter,
            "tol": self.tol,
            "random_state": self.random_state,
            "verbose": self.verbose,
            "sleep": self.sleep
            }
        
        params = {**parent_params, **child_params}
        return params

    def set_params(self, **params):
        rebuild_model = False
        for param, value in params.items():
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Parameter {param} does not exist for {self.__class__.__name__}")
            
            if rebuild_model:
                self.build_model()
            
    def build_model(self):
        self.model = ElasticNet(alpha = self.lmd,
                                l1_ratio = self.rho, max_iter = self.max_iter, tol = self.tol)
        
    def __compute_L(self, X_train):
        L = np.max(svd(X_train, compute_uv = False))**2
        return L
                
    def fit(self, X_train, y_train):
        if self.model is not None:
            self.model.fit(X_train, y_train)
        else:
            K = X_train.shape[1]
            X_train = np.array(X_train)
            N = len(y_train)
            y_train = np.array(y_train).reshape((N,1))
            L = self.__compute_L(X_train)
            gamma = 1 / L
            
            if self.cost_function == "huber":
                epsilon = self.compute_epsilon(X_train, y_train, self.huber_quantile)
                
            if self.random_state is not None:
                np.random.seed(self.random_state)
                theta = np.random.uniform(size=(K, 1))    
            else:
                theta = np.zeros((K,1))
                
            for m in np.arange(self.max_iter):
                theta_old = theta
                
                if self.cost_function == 'mse':
                    theta_bar = theta - gamma * self.grad_mse(X_train, y_train, theta)
                else:
                    theta_bar = theta - gamma * self.grad_huber(X_train, y_train, theta, epsilon)
                    
                theta_til = self.proximal(theta_bar, self.lmd, self.rho, gamma)
                theta = theta_til + (m / (m + 3)) * (theta_til - theta)
                gamma = gamma
                
                if self.verbose:
                    print(f'{m+1} iters finished.')
                    print(f'theta = {theta.T}')
                    print(f'Tolerance: {np.sum((theta-theta_old)**2)}')
                    if self.sleep is not None:
                        time.sleep(self.sleep)
                        
                if (np.sum((theta-theta_old)**2) < np.sum(theta_old**2*self.tol)) or (np.sum(np.abs(theta-theta_old))) == 0:
                    break
                
            self.theta = theta_old
        
        return self
    
    def predict(self, X):
        if self.model is not None:
            return self.model.predict(X).flatten()
            
        else:
            X = np.array(X)
            return (X @ self.theta).flatten()
        
    def score(self, X, y):
        y_pred = self.predict(X)
        r2 = r2_oos(y, y_pred)
        return r2
            




class GeneralizedLinearModelGroupLasso(OrdinaryLeastSquares):
    
    def __init__(self,
                 knots=3,
                 group_reg=1e-4,
                 l1_reg=1e-4,
                 random_state=12308,
                 *args,
                 **kwargs):
        self.knots = knots
        self.group_reg = group_reg
        self.random_state = random_state
        self.l1_reg = l1_reg
        super().__init__(*args, **kwargs)
        self.cost_function = "mse"
        self.__build_model()
        
    def get_params(self, deep = True):
        parent_params = super().get_params(deep)
        child_params = {
            "knots" : self.knots,
            "group_reg": self.group_reg,
            "l1_reg": self.l1_reg,
            "random_state": self.random_state
            }
        
        params = {**parent_params, **child_params}
        return params

    def set_params(self, **params):
        rebuild_model = False
        for param, value in params.items():
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Parameter {param} does not exist for {self.__class__.__name__}")
            
            if rebuild_model:
                self.__build_model()    

    def __SplineTransform(self, data, knots):
        spline_data = np.ones((data.shape[0], 1))
        
        for col in range(data.shape[-1]):
            i_dat = data[:, col]
            kbd = KBinsDiscretizer(n_bins=3, encode='onehot-dense', strategy='uniform')
            i_dat = i_dat.reshape(-1, 1)  # reshape to 2D array for compatibility with KBinsDiscretizer
            i_dum = kbd.fit_transform(i_dat)
            created_bins = i_dum.shape[1]
            
            if kbd.bin_edges_[0][0] in [-np.inf]:
                kbd.bin_edges_[0][0] = np.min(i_dat)
                        
            elif kbd.bin_edges_[0][-1] in [np.inf]:
                kbd.bin_edges_[0][-1] = np.max(i_dat)
                    
            if created_bins >= knots:
                for knot_i in range(knots):
                    i_dum[:,knot_i] = i_dum[:, knot_i] * ((i_dat.squeeze() - kbd.bin_edges_[0][knot_i])**2)
                    
            else:
                for bin_i in range(created_bins):
                    i_dum[:,bin_i] = i_dum[:, bin_i] * ((i_dat.squeeze() - kbd.bin_edges_[0][bin_i])**2)

                missing_bins = knots - created_bins
                missing_columns = np.zeros((i_dum.shape[0], missing_bins))
                i_dum = np.hstack((i_dum, missing_columns))

            spline_data = np.hstack((spline_data, i_dat, i_dum))

        return spline_data

    def __build_model(self):
        self.model = GroupLasso(
            groups = None, group_reg = self.group_reg, l1_reg = self.l1_reg,
            fit_intercept = False, random_state = self.random_state
        )

    def fit(self, X_train, y_train):
        groups = [[0]] + [[i] * (1 + self.knots) for i in range(1, X_train.shape[-1] + 1)]
        groups_flat = list(flatten(groups))
        self.model.set_params(groups = groups_flat)
        
        self.model.fit(self.__SplineTransform(X_train, self.knots), y_train)
    
    def predict(self,X_test):
        return self.model.predict(self.__SplineTransform(X_test, self.knots)).flatten()
    
    def score(self, X, y):
        y_pred = self.predict(X)
        r2 = self.r2_oos(y, y_pred)
        return r2
    




class RandomForest(Model):
    """
    RandomForest model for regression tasks, with support for scikit-learn and XGBoost implementations.

    Parameters
    ----------
    n_estimators : int, optional (default=300)
        The number of trees in the forest.

    max_depth : int, optional (default=None)
        The maximum depth of the tree.

    max_features : int, float, str or None, optional (default="auto")
        The number of features to consider when looking for the best split.

    method : str, optional (default="sklearn")
        The implementation method for RandomForest. Valid options are "sklearn" and "xgboost".

    n_jobs : int or None, optional (default=None)
        The number of jobs to run in parallel for both fit and predict.

    random_state : int, RandomState instance or None, optional (default=None)
        The random number seed to be used. If None, the random number generator is the RandomState instance used by np.random.

    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.

    set_params(**params)
        Set the parameters of the model.

    build_model()
        Build the RandomForest model based on the selected method.

    fit(X_train, y_train)
        Fit the model to the training data.

    predict(X_test)
        Predict the target values for the test data.

    score(X_test, y_test)
        Compute the out-of-sample R-squared score for the test data.
    """

    def __init__(self,
                 num_features,
                 n_estimators=300,
                 max_depth=None,
                 max_features="auto",
                 method="sklearn",
                 n_jobs=None,
                 random_state=None,
                 ):
        self.num_features = num_features
        self.n_estimators = n_estimators
        self.max_depth = max_depth
        self.max_features = max_features
        self.method = method
        self.n_jobs = n_jobs
        self.random_state = random_state

        if self.method not in ["sklearn", "xgboost"]:
            raise ValueError("Invalid value for parameter '{self.method}' specify either 'sklearn' or 'xgboost')
        if not ((isinstance(self.max_features, (int, float))) or (self.max_features in ["auto", "sqrt", "log"])):
            raise ValueError("Invalid value for parameter '{self.max_features}' specify either int, float, or the strings 'auto', 'sqrt', or 'log'")

        self.__build_model()

    def get_params(self, deep=True):
        return {
            "n_estimators": self.n_estimators,
            "max_depth": self.max_depth,
            "max_features": self.max_features,
            "method": self.method,
            "n_jobs": self.n_jobs,
            "random_state": self.random_state,
            "num_features": self.num_features
        }

    def set_params(self, **params):
        rebuild_model = False
        for param, value in params.items():
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{param}' for estimator {self.__class__.__name__}")
                
        if rebuild_model:
            self.build_model()

    def __build_model(self):
        if self.method == "sklearn":
            self.model = RandomForestRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                max_features=self.max_features,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )
        elif self.method == "xgboost":
            if self.max_features == "auto":
                self.model = XGBRFRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                colsample_bynode=1.0,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )
            elif self.max_features == "log":
                max_features = np.log(self.num_features)
                prop = max_features/self.num_features
                self.model = XGBRFRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                colsample_bynode=prop,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )
            elif self.max_features == "sqrt":
                max_features = np.sqrt(self.num_features)
                prop = max_features/self.num_features
                self.model = XGBRFRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                colsample_bynode=prop,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )
            elif isinstance(self.max_features, (int, np.int_)):
                prop = self.max_features/self.num_features
                self.model = XGBRFRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                colsample_bynode=prop,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )
            else:
                self.model = XGBRFRegressor(n_estimators=self.n_estimators,
                                                max_depth=self.max_depth,
                                                colsample_bynode=self.max_features,
                                                n_jobs=self.n_jobs,
                                                random_state=self.random_state
                                                )

    def fit(self, X_train, y_train):
        self.model.fit(X_train, y_train)
        return self

    def predict(self, X_test):
        return self.model.predict(X_test).ravel()

    def score(self, X_test, y_test):
        y_pred = self.predict(X_test)
        r2_oos = self.r2_oos(y_test, y_pred)
        return r2_oos





class GradientBoosting(Model):
    """
    Gradient Boosting model for regression tasks, with support for scikit-learn, CatBoost, and XGBoost implementations.
    
    This class enables to use different methods depending on the size of the dataset from the lowest to the largest
    
    Parameters
    ----------
    n_estimators : int, optional (default=100)
        The number of boosting stages (trees) to fit.
    
    max_depth : int, optional (default=1)
        The maximum depth of the individual regression estimators.
        
    learning_rate : float, optional (default=0.01)
        The learning rate, which controls the contribution of each tree in the model.
        
    loss_function : str, optional (default="huber")
        The loss function to be optimized. Valid options are "huber" and "mse".
        
    method : str, optional (default="catboost")
        The implementation method for Gradient Boosting. Valid options are "catboost", "sklearn", and "xgboost".
        
    random_state : int, RandomState instance or None, optional (default=None)
        The random number seed to be used. If None, the random number generator is the RandomState instance used by np.random.

    Methods
    -------
    get_params(deep=True)
        Get the parameters of the model.
        
    set_params(**params)
        Set the parameters of the model.
        
    build_model()
        Build the gradient boosting model based on the selected method.
        
    fit(X_train, y_train)
        Fit the model to the training data.
        
    predict(X_test)
        Predict the target values for the test data.
        
    score(X_test, y_test)
        Compute the out-of-sample R-squared score for the test data.
    """

    
    def __init__(self,
                 n_estimators=300,
                 max_depth=1,
                 learning_rate=0.01,
                 loss_function="huber",
                 method="catboost",
                 n_jobs=None,
                 verbose=None,
                 random_state=None
                 ):
        self.n_estimators = n_estimators
        self.max_depth = max_depth
        self.learning_rate = learning_rate
        self.loss_function = loss_function
        self.method = method
        self.n_jobs = n_jobs
        self.random_state = random_state
        self.verbose = verbose
        
        if self.loss_function not in ["huber", "mse"]:
            raise ValueError("Invalide value for parameter '{self.loss_function}' specify either 'huber' or 'mse'")
            
        if self.method not in ["sklearn", "catboost", "xgboost"]:
            raise ValueError("Invalide value for parameter '{self.method}' specify either 'sklearn', 'catboost', or 'xgboost")
            
        self.__build_model()
        
    def get_params(self, deep = True):
        return {
            "n_estimators": self.n_estimators,
            "max_depth": self.max_depth,
            "learning_rate": self.learning_rate,
            "loss_function": self.loss_function,
            "method": self.method,
            "n_jobs": self.n_jobs,
            "verbose" :self.verbose,
            "random_state": self.random_state
            }
    
    def set_params(self, **params):
        rebuild_model = False
        for param, value in params.items():
            if hasattr(self, param):
                setattr(self, param, value)
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{param}' for estimator {self.__class__.__name__}")
                
        if rebuild_model:
            self.__build_model()   

    def __build_model(self):
        if self.method == "catboost":
            if self.loss_function == "huber":
                self.model = CatBoostRegressor(iterations=self.n_estimators,
                                               learning_rate = self.learning_rate,
                                               depth = self.max_depth,
                                               loss_function = "Huber:delta=1",
                                               thread_count = self.n_jobs, 
                                               verbose = self.verbose,
                                               random_seed = self.random_state
                                               )
            else:
                self.model = CatBoostRegressor(iterations=self.n_estimators,
                                               learning_rate = self.learning_rate,
                                               depth = self.max_depth,
                                               loss_function = "MSE",
                                               thread_count = self.n_jobs,
                                               verbose = self.verbose,
                                               random_seed = self.random_state
                                               )
        elif self.method == "sklearn":
            if self.loss_function == "huber":
                if self.verbose in [True, 1]:
                    self.model = GradientBoostingRegressor(n_estimators = self.n_estimators, 
                                                           learning_rate = self.learning_rate,
                                                           max_depth = self.max_depth,
                                                           loss = "huber", alpha = 0.99,
                                                           verbose = 1,
                                                           random_state = self.random_state
                                                           )
                else:
                    self.model = GradientBoostingRegressor(n_estimators = self.n_estimators, 
                                                           learning_rate = self.learning_rate,
                                                           max_depth = self.max_depth,
                                                           loss = "huber", alpha = 0.99,
                                                           random_state = self.random_state
                                                           )
            else:
                if self.verbose in [True, 1]:
                    self.model = GradientBoostingRegressor(n_estimators = self.n_estimators, 
                                                           learning_rate = self.learning_rate,
                                                           max_depth = self.max_depth,
                                                           loss = "squared_error",
                                                           verbose = 1,
                                                           random_state = self.random_state
                                                           )
                else:
                    self.model = GradientBoostingRegressor(n_estimators = self.n_estimators, 
                                                           learning_rate = self.learning_rate,
                                                           max_depth = self.max_depth,
                                                           loss = "squared_error",
                                                           random_state = self.random_state
                                                           )
        else:
            if self.loss_function == "huber":
                if self.verbose in [True, 1]:
                    self.model = XGBRegressor(n_estimators = self.n_estimators,
                                                  learning_rate = self.learning_rate,
                                                  max_depth = self.max_depth,
                                                  objective = "reg:pseudohubererror",
                                                  n_jobs = self.n_jobs,
                                                  verbosity = 1,
                                                  random_state = self.random_state
                                                  )
                else:
                    self.model = XGBRegressor(n_estimators = self.n_estimators,
                                                  learning_rate = self.learning_rate,
                                                  max_depth = self.max_depth,
                                                  objective = "reg:pseudohubererror",
                                                  n_jobs = self.n_jobs,
                                                  verbosity = 0,
                                                  random_state = self.random_state
                                                  )
            else:
                if self.verbose in [True, 1]:
                    self.model = XGBRegressor(n_estimators = self.n_estimators,
                                                  learning_rate = self.learning_rate,
                                                  max_depth = self.max_depth,
                                                  objective = "reg:squarederror",
                                                  n_jobs = self.n_jobs,
                                                  verbosity = 1,
                                                  random_state = self.random_state
                                                  )
                else:
                    self.model = XGBRegressor(n_estimators = self.n_estimators,
                                                  learning_rate = self.learning_rate,
                                                  max_depth = self.max_depth,
                                                  objective = "reg:squarederror",
                                                  n_jobs = self.n_jobs,
                                                  verbosity = 0,
                                                  random_state = self.random_state
                                                  )
    
    def fit(self, X_train, y_train):
        self.model.fit(X_train, y_train)
        return self
        
    def predict(self, X_test):
        return self.model.predict(X_test).ravel()
        
    def score(self, X, y):
        y_pred = self.predict(X).ravel()
        r2_oos = self.r2_oos(y, y_pred)
        return r2_oos
        
        
        

class NeuralNetwork(Model):
    """
    A customizable neural network class built on top of the Keras Sequential API.

    Parameters
    ----------
    input_dim : int
        The dimension of the input data.
        
    n_layers : int, optional, default: 1
        The number of hidden layers in the neural network.
        
    learning_rate : float, optional, default: 0.001
        The learning rate for the Adam optimizer.
        
    l1_reg : float, optional, default: 1e-6
        The L1 regularization factor.
        
    epochs : int, optional, default: 100
        The number of training epochs.
        
    batch_size : int, optional, default: 200
        The batch size for training.
        
    batch_norm : bool, optional, default: True
        Whether to apply batch normalization between layers.
        
    random_state : int, optional, default: 1234
        The random seed for reproducibility.
        
    patience : int, optional, default: 6
        The number of epochs with no improvement after which training will be stopped (early stopping).
        
    early_stopping : bool, optional, default: True
        Whether to apply early stopping during training.
        
    dropout_rate : float, optional, default: None
        The dropout rate to apply after each layer. If both `dropout_rate` and `dropout_rates` are provided, an error will be raised.
        
    dropout_rates : list of float, optional, default: None
        A list of dropout rates for each layer. If both `dropout_rate` and `dropout_rates` are provided, an error will be raised.

    Attributes
    ----------
    nn : keras.Sequential
        The Keras Sequential model object.
        
    history : keras.callbacks.History
        The history object containing the training metrics.

    Methods
    -------
    r2_oos_nn(y_true, y_pred)
        Computes the out-of-sample R2 score.
        
    get_params(deep=True)
        Gets the parameters of the model.
        
    set_params(**params)
        Sets the parameters of the model.
        
    build_model()
        Builds the neural network model based on the given parameters.
        
    fit(X_train, y_train, validation_split=0.0)
        Fits the neural network model on the given training data.
    predict(X_test)
        Predicts the output for the given input data.
        
    score(X_test, y_test)
        Computes the out-of-sample R2 score on the test data.
        
    plot_training_history()
        Plots the training loss and R2 score.
        
    __str__()
        Returns a string representation of the neural network model.
    """
    
    
    def __init__(self,
                 input_dim,
                 n_layers=1,
                 learning_rate=0.001,
                 l1_reg=0.000001,
                 epochs=100,
                 batch_size=200,
                 batch_norm=True,
                 random_state=None,
                 patience=6, 
                 dropout_rate=None,
                 dropout_rates=None
                 ):
        self.input_dim = input_dim
        self.n_layers = n_layers
        self.learning_rate = learning_rate
        self.l1_reg = l1_reg
        self.epochs = epochs
        self.batch_size = batch_size
        self.batch_norm = batch_norm
        self.random_state = random_state
        self.patience = patience
        self.dropout_rate = dropout_rate
        self.dropout_rates = dropout_rates
        
        if self.random_state is not None:
            random.seed(self.random_state)
            np.random.seed(self.random_state)
            random.set_seed(self.random_state)
        
        if self.dropout_rate is not None and self.dropout_rates is not None:
            raise ValueError("Please provide either 'dropout_rate' or 'dropout_rates', not both.")
            
        self._build_model()

    @staticmethod
    def r2_oos_nn(y_true, y_pred):
        """
        Computes the out-of-sample R2 score based on Tensorflow requirements

        Parameters
        ----------
        y_true : array-like
            The true output values.
        y_pred : array-like
            The predicted output values.

        Returns
        -------
        r2 : float
            The out-of-sample R2 score.
        """
        
        SS_res = reduce_sum(square(subtract(y_true, y_pred)))
        SS_tot_without_mean = reduce_sum(square(y_true))
        r2 = subtract(1.0, divide(SS_res, SS_tot_without_mean))
        return r2
    
    def get_params(self, deep=True):        
        return {
            "input_dim": self.input_dim,
            "n_layers": self.n_layers,
            "learning_rate": self.learning_rate,
            "l1_reg": self.l1_reg,
            "epochs": self.epochs,
            "batch_size": self.batch_size,
            "batch_norm": self.batch_norm,
            "random_state": self.random_state,
            "patience": self.patience,
            "dropout_rate": self.dropout_rate,
            "dropout_rates":self.dropout_rates
        }
    
    def set_params(self, **params):        
        rebuild_model = False
        # Update the parameters based on the provided dictionary
        for key, value in params.items():
            if hasattr(self, key):
                setattr(self, key, value)
                # rebuild model only if parameters have been modified
                rebuild_model = True
            else:
                raise ValueError(f"Invalid parameter '{key}' for estimator {self.__class__.__name__}")
        if rebuild_model:
            self._build_model()

        
    def _build_model(self):
        """
        A protected method for building a neural network model based on the given parameters.

        The model is built as a sequential model with a specified number of hidden layers. Each hidden layer is followed
        by an optional batch normalization layer and an optional dropout layer, depending on the provided settings. The
        final output layer has a single neuron. The weights are initialized randomly and L1
        regularization is applied to the weights of all layers.
        """
  
        input_dim = self.input_dim
        max_hidden_layers = 5
        model = Sequential()
        model.add(Input(shape=(input_dim,)))          
        
        for layer in np.arange(self.n_layers, 0, -1):
            
            if self.n_layers > max_hidden_layers:
                model.add(Dense(2 ** layer, activation='relu', kernel_regularizer=l1(self.l1_reg)))
            else:
                model.add(Dense(2 ** (max_hidden_layers - (self.n_layers - layer)),
                                activation='relu', kernel_regularizer=l1(self.l1_reg)))
            if self.batch_norm:
                model.add(BatchNormalization())  
                
            if self.dropout_rates:
                model.add(Dropout(self.dropout_rates[self.n_layers - layer]))       
            elif self.dropout_rate is not None:
                model.add(Dropout(self.dropout_rate))           
                          
        model.add(Dense(1, kernel_regularizer=l1(self.l1_reg)))#, kernel_initializer = output_initializer))
        self.nn = model

    def fit(self, X_train, y_train):
        early_stopping = EarlyStopping(monitor = "loss", patience=self.patience)
        adam = Adam(learning_rate=self.learning_rate)
        self.nn.compile(optimizer=adam, loss="mse", metrics=[NeuralNetwork.r2_oos_nn])
        self.nn.fit(X_train, y_train, epochs=self.epochs, batch_size=self.batch_size, callbacks=[early_stopping])
        return self

    def predict(self, X_test):        
        # Check if the model has been trained
        if self.nn is None:
            raise ValueError("Model has not been trained. Call 'fit' before using 'predict'.")
            
        # Use the trained model to make predictions
        return self.nn.predict(X_test).flatten()

        
    def score(self, X, y):
        if self.nn is None:
            raise ValueError("Model has not been trained. Call 'fit' before using 'evaluate'.")
            
        y_pred = self.predict(X)
        r2_score = NeuralNetwork.r2_oos_nn(y, y_pred)
        return r2_score.numpy()
    
    def plot_training_history(self):
        """
        Plot the training loss and R2 Out-of-sample score
        """
        
        sns.set(style="whitegrid")  # Set the seaborn style for better aesthetics
        plt.figure(figsize=(14, 6))

        plt.subplot(1, 2, 1)
        plt.plot(self.history.history['loss'], label='Training Loss', color='blue', linewidth=2)
        plt.xlabel('Epoch')
        plt.ylabel('Loss')
        plt.legend()
        plt.title('Training Loss', fontsize=16)
        plt.suptitle('Training History', fontsize=20, y=1.05)  # Add a main title for both subplots

        plt.subplot(1, 2, 2)
        plt.plot(self.history.history['r2_oos_nn'], label='Training R2 Out-of-sample', color='green', linewidth=2)
        plt.xlabel('Epoch')
        plt.ylabel('R2')
        plt.legend()
        plt.title('Training R2 Out-of-sample', fontsize=16)

        plt.tight_layout()  # Adjust the layout for better spacing between subplots
        plt.show()

    def __str__(self):
        """
        Returns a summary of the neural network's architecture

        Returns
        -------
        str_repr : str
            The string representation of the model.
        """   
        
        if self.nn is not None:
            print(self.nn.summary())
            
        return ""
        
    


class EnsembleNeuralNetwork(NeuralNetwork):
    
    def __init__(self,
                 n_ensemble=10,
                 base_random_seed=None,
                 *args,
                 **kwargs
                 ):
        super().__init__(*args, **kwargs)
        self.n_ensemble = n_ensemble
        self.base_random_seed = base_random_seed
        self.__build_models()
        
    def get_params(self, deep = True):
        parent_params = super().get_params(deep)        
        child_params = {"n_ensemble": self.n_ensemble}
        all_params = {**parent_params, **child_params}
        return all_params
        
    def set_params(self, **params):
        rebuild_models = False
        for key, value in params.items():
            if hasattr(self, key):
                setattr(self, key, value)
                rebuild_models = True
            else:
                raise ValueError(f"Invalid parameter '{key}' for estimator {self.__class__.__name__}")
        if rebuild_models:
            self.build_models()
            
    def __build_models(self):
        self.models = []
        
        for i in range(self.n_ensemble):
            
            if self.base_random_seed is not None:
                self.random_state = self.base_random_seed + i
                random.seed(self.random_state)
                np.random.seed(self.random_state)
                set_seed(self.random_state)
                
            self._build_model()
            self.models.append(self.nn)

    def fit(self, X_train, y_train):
        
        for i, model in enumerate(self.models):
            print("\n", 35 * "#", f"Ensemble model: {i+1}/{self.n_ensemble}", 35 * "#")
            self.nn = model
            super().fit(X_train, y_train)
            print("\n")
            
        return self
            
    def predict(self, X_test):
        predictions = np.zeros((X_test.shape[0], self.n_ensemble))
        
        for i, model in enumerate(self.models):
            predictions[:, i] = model.predict(X_test).flatten()
            
        return predictions.mean(axis=1).flatten()


    def score(self, X_test, y_test):
        y_pred = self.predict(X_test)
        r2_score = NeuralNetwork.r2_oos_nn(y_test, y_pred)
        
        return r2_score.numpy()            
        
    def __str__(self):
        """
        Returns a summary of all models in the ensemble

        Returns
        -------
        str_repr : str
            The string representation of the models.
        """   
        
        for i, model in enumerate(self.models):
            print(65 * '-')
            print(f"Ensemble model: {i+1}/{len(self.models)}")
            print(65 * '-')
            print(model.summary())
            
        return ""    
            
            
            
            
           

###########################
# MODEL FITTING AND TUNNING
###########################


r2_oos = lambda actual, pred: Model.r2_oos(actual, pred)


def tune(model,
              param_grid,
              train,
              val,
              search="grid",
              random_state=1234,
              n_iter=10,
              categorical = None
              ):
    
    def objective(params, model, X_train, y_train, X_dev, y_dev):
        model.set_params(**params)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_dev)
        score = r2_oos(y_dev, y_pred)
        return -score  # Minimize the negative R^2
    
    
    X_train = np.array(train.iloc[:, 1:])
    y_train = np.array(train.iloc[:, 0])
    
    X_dev = np.array(val.iloc[:, 1:])
    y_dev = np.array(val.iloc[:, 0])
    param_list = {}
    
    best_score = - np.inf
    
    if search == "grid":
        param_list = list(ParameterGrid(param_grid))
        
    elif search == "random":
        param_list = list(ParameterSampler(param_grid, n_iter=n_iter, random_state=random_state))
        
    if search in ["grid", "random"]:
        for i, params in enumerate(param_list):
            print(" ")
            
            if search == "grid":
                print(f"Grid search: {i+1}/{len(param_list)}") 
            else:
                print(f"Random search: {i+1}/{len(param_list)}")  
                				
            print(60 * "-")    
            print(f"Hyperparameters tested: {params}")
            # Fit the model on the training data
            model.set_params(**params)
            model.fit(X_train, y_train)

            # Evaluate the model on the validation data
            y_pred = model.predict(X_dev)
            score = r2_oos(y_dev, y_pred)

            # Update the best parameters and score if necessary
            if score > best_score:
                best_params = params
                best_score = score
                
            print("Dev score: {:.2f} %".format(score * 100))
            print(60 * "-", "\n", "\n")    
            
        # Print the best hyperparameters and score
        print(" ")
        print(60 * "#")
        print("Best hyperparameters: ", best_params)
        print("Best dev score: {:.2f}".format(best_score * 100), "%")
        print(60 * "#", "\n", "\n")

        
    elif search == "bayesian":
        search_space = []
        
        if categorical is not None:
            param_non_cat = {key : [value[0], value[-1]] for key, value in param_grid.items() if key not in categorical}
            param_cat = {key : value for key, value in param_grid.items() if key in categorical}
            param_list = {**param_non_cat, **param_cat}
            
            for key in param_list.keys():
                if key not in categorical:
                    if isinstance(param_list[key][0], (float, np.float_)) and isinstance(param_list[key][1], (float, np.float_)):
                        if param_list[key][0] == param_list[key][-1]:
                            search_space.append(Categorical(param_list[key][0], name = key))
                            
                        else:
                            search_space.append(Real(param_list[key][0], param_list[key][1], name = key))

                    elif isinstance(param_list[key][0], (int, np.int_)) and isinstance(param_list[key][1], (int, np.int_)):
                        if param_list[key][0] == param_list[key][-1]:
                            search_space.append(Categorical(param_list[key][0], name = key))
                            
                        else:
                            search_space.append(Integer(param_list[key][0], param_list[key][1], name = key))
    
                    else:
                        raise ValueError(f"Invalid format for parameter {key}. Please use either float or int values")

                elif key in categorical:
                    search_space.append(Categorical(param_list[key], name = key))
                    
                else:
                    raise ValueError(f"Wrong name for categorical parameter {key}.")
                    print(f"Please use a valid name among those: {[key for key in param_grid.keys()]}")
            
        else:
            param_list = {key : [value[0], value[-1]] for key, value in param_grid.items()}
            
            for key in param_list.keys():
                if isinstance(param_list[key][0], (float, np.float_)) and isinstance(param_list[key][1], (float, np.float_)):
                    if param_list[key][0] == param_list[key][-1]:
                        search_space.append(Categorical(param_list[key][0], name = key))
                        
                    else:
                        search_space.append(Real(param_list[key][0], param_list[key][1], name = key))
                    
                elif isinstance(param_list[key][0], (int, np.int_)) and isinstance(param_list[key][1], (int, np.int_)):
                    if param_list[key][0] == param_list[key][-1]:
                        search_space.append(Categorical(param_list[key][0], name = key))
                        
                    else:
                        search_space.append(Integer(param_list[key][0], param_list[key][1], name = key))

                else:
                    raise ValueError(f"Invalid format for parameter {key}. Please use either float or int values")
                
                
        stopper = DeltaYStopper(delta=0.0001, n_best = 5)        
        optimizer = Optimizer(search_space, n_initial_points=5, n_jobs = -1)
        
        for iter_i in range(n_iter):
            print(" ")
            print(f"Bayesian search: {iter_i+1}/{n_iter}")

            x = optimizer.ask()
                
            params = {k: v for k, v in zip(param_list.keys(), x)}
            print( 60 * "-")    
            print(f"Hyperparameters tested: {params}")

            loss = objective(params, model, X_train, y_train, X_dev, y_dev)
            
            print("Dev score: {:.2f} %".format(-loss * 100))
            print(60 * "-", "\n", "\n")    

            optimizer.tell(x, loss)
            
            if iter_i > 0:
                if stopper(create_result(optimizer.Xi, optimizer.yi)):
                    print("No significative improvements of {:.2f} %.".format( - create_result(optimizer.Xi, optimizer.yi).fun * 100))
                    print("Stopping early")
                    print(" ")
                    break


        results = create_result(optimizer.Xi, optimizer.yi, optimizer.space, optimizer.rng, models=optimizer.models)
        best_params_list = results.x        
        best_score = - results.fun
        # Create a dictionary to store the best hyperparameters
        best_params = {key: value for key, value in  zip(param_list.keys(), best_params_list)}
        
        print(" ")       
        print(60 * "#")
        print(f"Best hyperparameters: {best_params}")
        print("Best dev score: {:.2f}".format(best_score * 100), "%")
        print(60 * "#", "\n", "\n")
        
    else:
        raise ValueError(f"{search} wrong value for parameter method", "\n")
        print("Specify either 'grid', 'random' or 'bayesian' for method parameter")


    model.set_params(**best_params)
        
    X_train_full = np.concatenate((X_train, X_dev))
    y_train_full = np.concatenate((y_train, y_dev))
        
    model.fit(X_train_full, y_train_full)
    
    return model





def increment_data(train, val, test):

    new_train_end = train.index.get_level_values(0)[-1] + pd.DateOffset(years=1)
    new_val_end = val.index.get_level_values(0)[-1] + pd.DateOffset(months=12)
    new_test_start = new_val_end + pd.DateOffset(months=1)
    
    new_train = df_new.loc[:new_train_end]
    new_val = df_new.loc[new_train_end + pd.DateOffset(months=1):new_val_end]
    new_test = df_new.loc[new_test_start:new_test_start + pd.DateOffset(months = 11)]
        
    return new_train, new_val, new_test





def run_model(model, data, param_grid = None, **kwargs):

    train = df_new.loc[:'1974-12-31']
    val = df_new.loc['1975':'1986-12-31']
    test = df_new.loc['1987']
    
    end_date = data.index.get_level_values(0)[-1]
    filename = f'predictions_{model}/predictions_{model}.xlsx'
    filename_param = f'predictions_{model}/parameters_{model}.xlsx'
    
    i = 2
    while os.path.isfile(filename):
        filename = f'predictions_{model}/predictions_{model}_{i}.xlsx'
        if model not in ["ols", "ols3"]:
            filename_param = f'predictions_{model}/parameters_{model}_{i}.xlsx'
        i += 1
            
    pd.DataFrame().to_excel(filename) 
    if model not in ["ols", "ols3"]:
        pd.DataFrame().to_excel(filename_param)
        
    y_pred = np.array([])
    stop_flag = False

    if model in ["ols", "ols3"]:        
        if model == "ols3":
            features_3 = ["Excess Return",'mvel1','bm','mom1m','mom6m','mom12m','mom36m']
            train = train.loc[:, features_3]
            val = val.loc[:, features_3]
            test = test.loc[:, features_3]
                    
        try:
            
            while not stop_flag and val.index.get_level_values('date')[-1] < end_date:
                start_fit = time.time()
                # Debug:
                new_train_start = train.index.get_level_values(0)[0].strftime('%Y-%m-%d')
                new_train_end = train.index.get_level_values(0)[-1].strftime('%Y-%m-%d') 
                new_val_start = val.index.get_level_values(0)[0].strftime('%Y-%m-%d')
                new_val_end = val.index.get_level_values(0)[-1].strftime('%Y-%m-%d') 
                new_test_start = test.index.get_level_values(0)[0].strftime("%Y-%m-%d")
                new_test_end = test.index.get_level_values(0)[-1].strftime("%Y-%m-%d")
                year_test_start = new_test_start[:4]
                
                print(" ")
                print(30 * "#")
                if model == "ols": print(10 * " ", "Model OLS")
                else: print(10 * " ", "Model OLS-3")
                
                print(30 * "#")
                print('Train:', new_train_start, ":", new_train_end)
                print('Val:  ',  new_val_start, ":", new_val_end)
                print('Test: ',  new_test_start, ":", new_test_end, "\n")
                
                X_train = np.array(train.iloc[:, 1:])
                y_train = np.array(train.iloc[:, 0])
                
                X_dev = np.array(val.iloc[:, 1:])
                y_dev = np.array(val.iloc[:, 0])
                
                X_test = np.array(test.iloc[:, 1:])
                y_test = np.array(test.iloc[:, 0])

                X_train_full = np.concatenate((X_train, X_dev))
                y_train_full = np.concatenate((y_train, y_dev))

                ols = OrdinaryLeastSquares(cost_function="huber")
                ols.fit(X_train_full, y_train_full)
                
                y_pred = ols.predict(X_test)
                score = ols.score(X_test, y_test)
                
                print('R2 OOS {}: {:.2f} %'.format(year_test_start, score * 100), "\n")

                                                
                if model == "ols":
                    pred = pd.DataFrame({'Pred OLS': y_pred}, index = test.index.get_level_values(0).strftime("%Y-%m-%d"))
                    
                    with pd.ExcelWriter(filename, mode = "a", engine = "openpyxl") as writer:    
                        pred.to_excel(writer, index=True, sheet_name=f'Predictions OLS {year_test_start}')
                else:
                    pred = pd.DataFrame({'Pred OLS-3': y_pred}, index = test.index.get_level_values(0).strftime("%Y-%m-%d"))
                    
                    with pd.ExcelWriter(filename, mode = "a", engine = "openpyxl") as writer:
                        pred.to_excel(writer, index=True, sheet_name=f'Predictions OLS-3 {year_test_start}')
                 
                end_fit = time.time()
                time_fit = end_fit - start_fit
                time_fit_min = time_fit / 60

                print("It took {:.2f} seconds, or {:.2f} minutes for {} testing date".format(time_fit, time_fit_min, year_test_start), "\n", "\n", "\n")
                
                train, val, test = increment_data(train, val, test)
                
                if val.index.get_level_values('date')[-1] == end_date:
                    break
                
                
        except KeyboardInterrupt:
            print("Program stopped")
            print(f"Last testing date: {year_test_start}")
            sys.exit(1)
            
        except IOError as e:
            print(f"An I/O error occured: {str(e)} ")
            sys.exit(1)
            
        except Exception as e:
            stop_flag = True
            print(f"An exception occurred: {str(e)}")
            sys.exit(1)
            
            
        workbook = openpyxl.load_workbook(filename)
        first_sheet_name = workbook.sheetnames[0]
        first_sheet = workbook[first_sheet_name]
        workbook.remove(first_sheet)
        workbook.save(filename)
        
        print(" ")
        print(50 * "#")        
        print("Fitting process finished succesfully for {} model".format(model.upper()))
        print(50 * "#")    


            
    else:    
        
        try:
            
            base_model = None
            best_model = None
            best_params = None
            
            while not stop_flag and val.index.get_level_values('date')[-1] < end_date:
                start_fit = time.time()
                # Debug:
                new_train_start = train.index.get_level_values(0)[0].strftime('%Y-%m-%d')
                new_train_end = train.index.get_level_values(0)[-1].strftime('%Y-%m-%d') 
                new_val_start = val.index.get_level_values(0)[0].strftime('%Y-%m-%d')
                new_val_end = val.index.get_level_values(0)[-1].strftime('%Y-%m-%d') 
                new_test_start = test.index.get_level_values(0)[0].strftime("%Y-%m-%d")
                new_test_end = test.index.get_level_values(0)[-1].strftime("%Y-%m-%d")
                year_test_start = new_test_start[:4]
                
                print(" ")
                print(30 * "#")
                print(10 * " ", "Model {}".format(model.upper()))
                print(30 * "#")
                print('Train:', new_train_start, ":", new_train_end)
                print('Val:  ',  new_val_start, ":", new_val_end)
                print('Test: ',  new_test_start, ":", new_test_end, "\n")
                
                if model == "pls":
                    base_model = PartialLeastSquaresRegression()
                    best_model = tune(base_model, param_grid, train, val, search = "grid")
                    
                if model == "pcr":
                    base_model = PrincipalComponentRegression()
                    best_model = tune(base_model, param_grid, train, val, search = "grid")
                    
                if model == "enet":
                    base_model = ElasticNetRegression(cost_function = "mse")
                    best_model = tune(base_model, param_grid, train, val,
                                           search= "bayesian", n_iter = 30)
                if model == "glm":
                    base_model = GeneralizedLinearModelGroupLasso()
                    best_model = tune(base_model, param_grid, train, val, search = "grid")

                if model == "rf": 
                    base_model = RandomForest(num_features=train.shape[-1], n_jobs = -1, method = "xgboost")
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 6)
                    
                if model == "gb":
                    base_model = GradientBoosting(n_jobs = -1, verbose = False, method = "xgboost")
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 6, categorical = ["max_depth"])
                                        
                if model == "nn1":
                    base_model = EnsembleNeuralNetwork(input_dim = (train.shape[-1] - 1), n_layers = 1,
                                                       batch_size = 200, base_random_seed = 1234)
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 15, categorical = ["dropout_rate"])
                                        
                if model == "nn2":
                    base_model = EnsembleNeuralNetwork(input_dim = (train.shape[-1] - 1), n_layers = 2, base_random_seed = 1234)
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 15, categorical = ["dropout_rate"])
                    
                if model == "nn3":
                    base_model = EnsembleNeuralNetwork(input_dim = (train.shape[-1] - 1), n_layers = 3, base_random_seed = 1234)
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 15, categorical = ["dropout_rate"])
                    
                if model == "nn4":
                    base_model = EnsembleNeuralNetwork(input_dim = (train.shape[-1] - 1), n_layers = 4, base_random_seed = 1234)
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 15, categorical = ["dropout_rate"])
                    
                if model == "nn5":
                    base_model = EnsembleNeuralNetwork(input_dim = (train.shape[-1] - 1), n_layers = 5, base_random_seed = 1234)
                    best_model = tune(base_model, param_grid, train, val,
                                           search = "bayesian", n_iter = 15, categorical = ["dropout_rate"])
                    
                    
                X_test = np.array(test.iloc[:, 1:])
                y_test = np.array(test.iloc[:, 0])
                    
                y_pred = best_model.predict(X_test)
                score = best_model.score(X_test, y_test)
                
                print("R2 OOS {}: {:.2f} %".format(year_test_start, score * 100), "\n")

                pred = pd.DataFrame({'Pred {}'.format(model.upper()): y_pred}, index = test.index.get_level_values(0).strftime("%Y-%m-%d"))
                with pd.ExcelWriter(filename, mode = "a", engine = "openpyxl") as writer:
                    pred.to_excel(writer, index=True, sheet_name='Predictions {} {}'.format(model.upper(), year_test_start))
                print(best_params)    
                param = pd.DataFrame(best_params, index = [year_test_start])
                with pd.ExcelWriter(filename_param, mode = "a", engine = "openpyxl") as writer:
                    param.to_excel(writer, index=True, sheet_name='Hyperparams {} {}'.format(model.upper(), year_test_start))

                    
                end_fit = time.time()
                time_fit = end_fit - start_fit
                time_fit_min = time_fit / 60

                print("It took {:.2f} seconds, or {:.2f} minutes for {} testing date".format(time_fit, time_fit_min, year_test_start), "\n", "\n", "\n")

                train, val, test = increment_data(train, val, test)
                
                if val.index.get_level_values(0)[-1] == end_date:
                    break
                                
        except KeyboardInterrupt:
            print("Program stopped")
            print(f"Last testing date: {year_test_start}")
            sys.exit(1)
            
        except IOError as e:
            print(f"An I/O error occured: {str(e)} ")
            print(f"Last testing date: {year_test_start}")
            sys.exit(1)

        except Exception as e:
            stop_flag = True
            print(f"An exception occurred: {str(e)}")
            print(f"Last testing date: {year_test_start}")
            sys.exit(1)
            
        # Remove the first blank sheet for the predictions   
        workbook = openpyxl.load_workbook(filename)
        first_sheet_name = workbook.sheetnames[0]
        first_sheet = workbook[first_sheet_name]
        workbook.remove(first_sheet)
        workbook.save(filename)
        
        # Remove the first blank sheet for the hyperparameters  
        workbook = openpyxl.load_workbook(filename)
        first_sheet_name = workbook.sheetnames[0]
        first_sheet = workbook[first_sheet_name]
        workbook.remove(first_sheet)
        workbook.save(filename)

        
        print(" ")
        print(" ")
        print(50 * "#")        
        print("Fitting process finished succesfully for {} model".format(model.upper()))
        print(50 * "#")    


            
   

def run_models(data, model_configs):
    
    param_grid_nn = {"l1_reg": np.linspace(1e-5, 1e-3, 10000),
                     "learning_rate": np.linspace(1e-3, 1e-2, 10000),
                     "dropout_rate": [None, 0.05, 0.1, 0.15, 0.2]
                     }

    default_model_configs = {
        "ols": {"run": False, "param_grid": None},
        
        "ols3": {"run": False, "param_grid": None},
        
        "pls": {"run": False, "param_grid": {"n_components": np.arange(1, 6 + 1)}},
        
        "pcr": {"run": False, "param_grid": {"n_components": [1, 2, 3, 4, 5, 8, 10, 12, 15, 20]}},
        
        "enet": {"run": False, "param_grid": {"lmd": np.linspace(1e-4, 1e-1, 1000)}},
        
        "glm": {"run": False, "param_grid": {"group_reg": [1e-5, 1e-4, 1e-3, 1e-2],
                                            "l1_reg": [1e-5, 1e-4, 1e-3, 1e-2]}},
        
        "rf": {"run": False, "param_grid": {"max_depth": np.arange(1, 6 + 1),
                                            "max_features": [3, 5, 8, 10, 12, 15, 20, 30, 40, 50, 80]}},
        
        "gb": {"run": False, "param_grid": {"n_estimators": np.arange(20, 1000),
                                            "max_depth": [1, 2],
                                            "learning_rate": np.linspace(0.01, 0.1, 10000)}},
        
        "nn1": {"run": False, "param_grid": param_grid_nn},
        
        "nn2": {"run": False, "param_grid": param_grid_nn},
        
        "nn3": {"run": False, "param_grid": param_grid_nn},
        
        "nn4": {"run": False, "param_grid": param_grid_nn},
        
        "nn5": {"run": False, "param_grid": param_grid_nn}
    }
    
    # Update default configs with user-specified configs
    for model_name, config in model_configs.items():
        default_model_configs[model_name]["run"] = config["run"]
        
    start_models = time.time()
    
    for model_name, config in default_model_configs.items():
        if config["run"]:
            print("Running", model_name.upper(), "...")
            
            start_model = time.time()
            run_model(model_name, data, param_grid=config["param_grid"])
            
            end_model = time.time()
            time_model = end_model - start_model
            time_model_min = time_model / 60
            time_model_hour = time_model_min / 60
            
            print(" ")
            print("It took {:.2f} seconds, or {:.2f} minutes, or {:.2f} hours to run {} model".format(time_model, time_model_min, time_model_hour, model_name.upper()), "\n", "\n", "\n", "\n")
            
            
    end_models = time.time()
    time_models = end_models - start_models
    time_models_min = time_models / 60
    time_models_hour = time_models_min / 60
    
    print(71 * "%")
    print("It took {:.2f} seconds, or {:.2f} minutes, or {:.2f} hours to run all models".format(time_models, time_models_min, time_models_hour))
    print(71 * "%", "\n", "\n", "\n")
        
        
    

args = parser.parse_args()

ols = args.run_ols
ols3 = args.run_ols3
pls = args.run_pls
pcr = args.run_pcr
enet = args.run_enet
glm = args.run_glm
rf = args.run_rf
gb = args.run_gb
nn1 = args.run_nn1
nn2 = args.run_nn2
nn3 = args.run_nn3
nn4 = args.run_nn4
nn5 = args.run_nn5

model_configs = {
    
    "ols": {"run": ols},
    "ols3": {"run": ols3},
    "pls": {"run": pls},
    "pcr": {"run": pcr},
    "enet": {"run": enet},
    "glm": {"run": glm},
    "rf": {"run": rf},
    "gb": {"run": gb},
    "nn1": {"run": nn1},
    "nn2": {"run": nn2},
    "nn3": {"run": nn3},
    "nn4": {"run": nn4},
    "nn5": {"run": nn5}
    
}


run_models(df_new, model_configs)


sys.exit(1)



